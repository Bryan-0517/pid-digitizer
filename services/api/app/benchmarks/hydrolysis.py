from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.models import EngineeringConnection, EngineeringEntity, EngineeringGraph, GraphMetadata

DOCUMENT_ID = "benchmark:hydrolysis"
UNASSIGNED_PAGE_ID = "benchmark:hydrolysis:unassigned"
FIXTURE_TIMESTAMP = "1970-01-01T00:00:00Z"
WORKBOOK_GLOB = "*.xlsx"
SCREEN_PATTERN = re.compile(r"IMG_\d+\.(?:JPG|JPEG|PNG)", re.IGNORECASE)


@dataclass(frozen=True)
class ConversionResult:
    graph: EngineeringGraph
    report: dict[str, Any]


def convert_workbook(workbook_path: Path) -> ConversionResult:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    equipment_rows = _records(workbook.worksheets[4])
    connection_rows = _records(workbook.worksheets[5])
    instrument_rows = _records(workbook.worksheets[6])

    entities: list[EngineeringEntity] = []
    node_counts = Counter(_required(row, 0) for row in equipment_rows)
    node_occurrences: Counter[str] = Counter()
    source_nodes: dict[str, list[str]] = {}
    duplicate_source_nodes: list[dict[str, Any]] = []
    for row in equipment_rows:
        source_id = _required(row, 0)
        node_occurrences[source_id] += 1
        suffix = ""
        if node_counts[source_id] > 1:
            suffix = f":{_value(row, 1) or 'unknown'}:{node_occurrences[source_id]}"
            duplicate_source_nodes.append({
                "sourceNodeId": source_id,
                "areaId": _value(row, 1),
                "occurrence": node_occurrences[source_id],
            })
        canonical_id = f"hydrolysis:equipment:{source_id}{suffix}"
        source_nodes.setdefault(source_id, []).append(canonical_id)
        properties = _properties(row, {
            0: "sourceNodeId", 1: "areaId", 4: "sourceEquipmentType", 5: "systemSequence",
            6: "upstreamSummary", 7: "downstreamSummary", 8: "sourceScreens",
            9: "extractionBasis", 11: "sourceStatus", 12: "remarks",
        })
        entities.append(EngineeringEntity.model_validate({
            "id": canonical_id,
            "documentId": DOCUMENT_ID,
            "pageId": UNASSIGNED_PAGE_ID,
            "kind": "boundary" if source_id.startswith("BND_") else "equipment",
            "subtype": _value(row, 4),
            "tag": _value(row, 2),
            "displayName": _value(row, 3),
            "properties": properties,
            "confidence": _confidence(row, 10),
            "assertion": _assertion(_value(row, 9), _value(row, 11)),
            "provenance": _provenance(workbook_path.name, "equipment_nodes", source_id, row, 8, 9, 10, 11, 12),
            "createdAt": FIXTURE_TIMESTAMP,
            "updatedAt": FIXTURE_TIMESTAMP,
        }))

    instrument_ids: dict[str, str] = {}
    for row in instrument_rows:
        source_id = _required(row, 0)
        canonical_id = f"hydrolysis:instrument:{source_id}"
        instrument_ids[source_id] = canonical_id
        entities.append(EngineeringEntity.model_validate({
            "id": canonical_id,
            "documentId": DOCUMENT_ID,
            "pageId": UNASSIGNED_PAGE_ID,
            "kind": "instrument",
            "tag": _value(row, 2),
            "displayName": _value(row, 3),
            "properties": _properties(row, {
                0: "sourceInstrumentId", 1: "areaId", 3: "variableName", 4: "ownerNodeId",
                5: "unit", 6: "role", 7: "sourceScreens", 8: "extractionBasis",
                10: "sourceStatus", 11: "remarks",
            }),
            "confidence": _confidence(row, 9),
            "assertion": _assertion(_value(row, 8), _value(row, 10)),
            "provenance": _provenance(workbook_path.name, "instrument_register", source_id, row, 7, 8, 9, 10, 11),
            "createdAt": FIXTURE_TIMESTAMP,
            "updatedAt": FIXTURE_TIMESTAMP,
        }))

    connections: list[EngineeringConnection] = []
    broken_connections: list[dict[str, Any]] = []
    unsupported: list[dict[str, Any]] = []
    for row in connection_rows:
        source_id = _required(row, 0)
        from_id, to_id = _value(row, 2), _value(row, 3)
        missing = [value for value in (from_id, to_id) if value not in source_nodes]
        ambiguous = [value for value in (from_id, to_id) if len(source_nodes.get(value, [])) > 1]
        if missing or ambiguous:
            broken_connections.append({
                "sourceConnectionId": source_id,
                "missingNodeIds": missing,
                "ambiguousNodeIds": ambiguous,
            })
            continue
        if from_id == to_id:
            unsupported.append({"register": "process_connections", "sourceId": source_id, "reason": "self-loop not explicitly authorized by source"})
            continue
        source_type = _value(row, 5)
        kind = {"Process": "process", "Utility": "utility"}.get(source_type, "unknown")
        connections.append(EngineeringConnection.model_validate({
            "id": f"hydrolysis:process:{source_id}",
            "documentId": DOCUMENT_ID,
            "sourceEntityId": source_nodes[from_id][0],
            "targetEntityId": source_nodes[to_id][0],
            "kind": kind,
            "medium": _value(row, 4),
            "direction": "source_to_target",
            "properties": _properties(row, {
                0: "sourceConnectionId", 1: "areaId", 2: "sourceFromNodeId",
                3: "sourceToNodeId", 5: "sourceConnectionType",
                6: "sourceScreens", 7: "extractionBasis", 9: "sourceStatus", 10: "notes",
            }),
            "confidence": _confidence(row, 8),
            "assertion": _assertion(_value(row, 7), _value(row, 9)),
            "provenance": _provenance(workbook_path.name, "process_connections", source_id, row, 6, 7, 8, 9, 10),
            "createdAt": FIXTURE_TIMESTAMP,
            "updatedAt": FIXTURE_TIMESTAMP,
        }))

    broken_owners: list[dict[str, Any]] = []
    missing_owners: list[dict[str, Any]] = []
    ownership_count = 0
    for row in instrument_rows:
        instrument_id = _required(row, 0)
        owner_id = _value(row, 4)
        if not owner_id:
            missing_owners.append({"sourceInstrumentId": instrument_id})
            continue
        if owner_id not in source_nodes:
            broken_owners.append({"sourceInstrumentId": instrument_id, "ownerNodeId": owner_id})
            continue
        if len(source_nodes[owner_id]) > 1:
            broken_owners.append({
                "sourceInstrumentId": instrument_id,
                "ownerNodeId": owner_id,
                "reason": "ambiguous duplicate source node ID",
            })
            continue
        assertion = _assertion(_value(row, 8), _value(row, 10))
        connections.append(EngineeringConnection.model_validate({
            "id": f"hydrolysis:ownership:{instrument_id}",
            "documentId": DOCUMENT_ID,
            "sourceEntityId": instrument_ids[instrument_id],
            "targetEntityId": source_nodes[owner_id][0],
            "kind": "ownership",
            "direction": "source_to_target",
            "properties": {"sourceInstrumentId": instrument_id, "sourceOwnerNodeId": owner_id},
            "confidence": _confidence(row, 9),
            "assertion": assertion,
            "provenance": _provenance(workbook_path.name, "instrument_register", instrument_id, row, 7, 8, 9, 10, 11),
            "createdAt": FIXTURE_TIMESTAMP,
            "updatedAt": FIXTURE_TIMESTAMP,
        }))
        ownership_count += 1

    tags = [entity.tag for entity in entities if entity.tag]
    duplicate_tags = [
        {"tag": tag, "count": count}
        for tag, count in sorted(Counter(tags).items()) if count > 1
    ]
    graph = EngineeringGraph(
        schema_version="0.1", document_id=DOCUMENT_ID,
        entities=entities, connections=connections,
        metadata=GraphMetadata(
            name="Hydrolysis pre-DEXPI benchmark",
            description="Deterministic conversion of uncertain DCS reference registers; not certified engineering truth.",
            source_kind="dcs",
        ),
    )
    report = {
        "documentId": DOCUMENT_ID,
        "entityCount": len(entities),
        "connectionCount": len(connections),
        "equipmentNodeCount": len(equipment_rows),
        "instrumentCount": len(instrument_rows),
        "ownershipConnectionsCreated": ownership_count,
        "duplicateTagWarnings": duplicate_tags,
        "duplicateSourceNodeWarnings": duplicate_source_nodes,
        "brokenSourceTargetReferences": broken_connections,
        "brokenInstrumentOwnerReferences": broken_owners,
        "missingInstrumentOwners": missing_owners,
        "unsupportedUnmappedSourceRecords": unsupported,
        "validationWarnings": [
            "All pageId values use the T007 unassigned sentinel; source screens remain in provenance.",
            "No geometry was imported because the source registers contain no verified normalized geometry.",
        ],
        "validationErrors": [],
    }
    return ConversionResult(graph=graph, report=report)


def write_conversion(result: ConversionResult, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_json(output_dir / "engineering_graph.json", result.graph.model_dump(by_alias=True, exclude_none=True))
    _write_json(output_dir / "import_report.json", result.report)


def _records(sheet: Any) -> list[tuple[Any, ...]]:
    return [tuple(_repair(value) for value in row) for row in list(sheet.iter_rows(values_only=True))[4:] if any(value is not None for value in row)]


def _repair(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    output: list[str] = []
    run: list[str] = []
    def flush() -> None:
        if not run:
            return
        text = "".join(run)
        try:
            candidate = text.encode("latin1").decode("gb18030")
            cjk_count = sum("\u3400" <= character <= "\u9fff" for character in candidate)
            output.append(candidate if cjk_count >= 2 else text)
        except UnicodeError:
            output.append(text)
        run.clear()
    for character in value:
        if ord(character) <= 255:
            run.append(character)
        else:
            flush()
            output.append(character)
    flush()
    return "".join(output).strip()


def _value(row: tuple[Any, ...], index: int) -> Any:
    value = row[index] if index < len(row) else None
    return None if value in (None, "", "—") else value


def _required(row: tuple[Any, ...], index: int) -> str:
    value = _value(row, index)
    if not isinstance(value, str):
        raise ValueError(f"required source ID missing at column {index}")
    return value


def _confidence(row: tuple[Any, ...], index: int) -> float | None:
    value = _value(row, index)
    return float(value) if value is not None else None


def _assertion(basis: str | None, status: str | None) -> dict[str, str]:
    evidence = basis or ""
    mode = "inferred" if any(marker in evidence for marker in ("推断", "AI", "流程")) else "observed"
    if status == "已确认":
        review_status = "confirmed"
    elif status in {"待现场核实", "待补资料", "待确认"}:
        review_status = "needs_source"
    else:
        review_status = "unreviewed"
    return {"mode": mode, "reviewStatus": review_status}


def _properties(row: tuple[Any, ...], mapping: dict[int, str]) -> dict[str, Any]:
    return {name: value for index, name in mapping.items() if (value := _value(row, index)) is not None}


def _provenance(
    workbook_name: str, register: str, source_id: str, row: tuple[Any, ...],
    screen_index: int, basis_index: int, confidence_index: int, status_index: int,
    note_index: int,
) -> list[dict[str, Any]]:
    confidence = _confidence(row, confidence_index)
    status, basis, note = _value(row, status_index), _value(row, basis_index), _value(row, note_index)
    summary = "; ".join(f"{label}: {value}" for label, value in (("basis", basis), ("status", status), ("note", note)) if value)
    evidence: list[dict[str, Any]] = [{
        "id": f"hydrolysis:evidence:{register}:{source_id}",
        "sourceType": "spreadsheet",
        "sourceRef": f"{workbook_name}#{register}:{source_id}",
        "note": summary or None,
        "confidence": confidence,
    }]
    screens = sorted(set(SCREEN_PATTERN.findall(str(_value(row, screen_index) or ""))))
    evidence.extend({
        "id": f"hydrolysis:evidence:{register}:{source_id}:screen:{screen.lower()}",
        "sourceType": "page_image", "sourceRef": screen,
        "note": f"Source screen reference from {register}; no T007 page assignment.",
        "confidence": confidence,
    } for screen in screens)
    return evidence


def _write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
